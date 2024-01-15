#Ultima actualización: 28/12/2023

from exchangelib import Credentials, Account, DELEGATE, Configuration,IMPERSONATION,FileAttachment, Credentials, Configuration, Account, DELEGATE
from datetime import datetime
import warnings
from tabulate import tabulate
warnings.filterwarnings("ignore")
from exchangelib.attachments import FileAttachment
import tkinter as tk
from tkinter import ttk
import re
import pandas as pd
import PySimpleGUI as sg
from tkinter import Tk, Button, messagebox, Label
from tkcalendar import Calendar, DateEntry
from dateutil import parser
from datetime import timedelta
import datetime
import sys
import time
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl import load_workbook
from getpass import getpass
from babel import numbers


global df
now = datetime.datetime.now()

# Crear el DataFrame
df = pd.DataFrame(columns=["Codi", "Aplicació", "Tecnologia", "Resultat", "Urgent", "Té incidència", "Incidencia associada", "Observacions", "Data"])

# Define las categorías y entornos
categories = {
    "Websphere": [".w61"],
    ".NET": [".NET"],
    "Client/Servidor": [".NT"],
    "BIM": [".BIM"],
    "BI": [".BI"],
    "Documentum": [".doc", ".wdk"],
    "Portlet": [".plr"],
    "Devops": [
        "Publicació d'API", "Petició desplegament DevOps",
        "DevOps", "Petició desplegament/execució scripts",
        "Petició desplegament/subscripció",
        "Petició desplegament/publicació d'API",
        "Petició de creació d'esquema BD",
        "Petició desplegament/creació esquema"
    ],
    "Cognos": [".CBI", ".CDM"],
    "Paquet": ["Fi Distribució tècnica paquet"],
    "BBDD": [".BD", "Scripts+Normal"],
    "Pegats": ["Distribucio pegats seguretat","Distribució pegats seguretat"],
    "Otros": ["SILTRA", "TREM", "FICOR"]
}

entornos = {
    "pre": [
        "Munteu la maqueta", "Homologar", "Muntar la maqueta",
        "Muntar maqueta", "Assignació d'Assistència",
        "Petició desplegament genèric SIA a PRE",
        "Detindre aplicació", "RE:", "Ok detenció aplicació a PRO",
        "Desinstal·lació aplicació", "RE:","pendent de distribució","RV:","Re:","Incidència"
    ]
}



def extract_dates_from_text(text):
    try: 
        date_formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d-%m-%y", "%d/%m/%y"]

        for date_format in date_formats:
            if text:
                dates = re.findall(r"\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}", text)
                for date_str in dates:
                    try:
                        #date = datetime.strptime(date_str, date_format).date() INVESTIGAR PORQUE NO FUNCIONA
                        date = datetime.datetime.strptime(date_str, date_format).strftime("%Y-%m-%d")
                        return date
                    except ValueError:
                        pass
        return None
    except:
        sg.Popup("Error al extraer las fechas del correo con asunto: " + text.split(":",1)[1])
        return None

def get_email_parts(email):
    if isinstance(email, FileAttachment):
        return []
    email_parts = [email]  
    email_parts.extend(attachment for attachment in email.attachments if attachment.is_inline)
    return email_parts

def get_folder_emails_recursive(folder,fechaInicio=None, fechaFinal=None):
    try:
        emails = []
        for email in folder.all():
            if isinstance(email, FileAttachment):
                continue

            if fechaInicio and fechaFinal:
                date = extract_dates_from_text(email.text_body)
                if date:
                    date = datetime.datetime.strptime(date, "%Y-%m-%d")
                    if date and fechaInicio <= date <= fechaFinal:
                        df = classify_email(email, date)
            else:
                date = extract_dates_from_text(email.text_body)
                if date:
                    date = datetime.datetime.strptime(date, "%Y-%m-%d")
                    df = classify_email(email, date)

        for subfolder in folder.children:
            emails.extend(get_folder_emails_recursive(subfolder, fechaInicio, fechaFinal))

        return emails
    except:
        sg.Popup("Fallo al obtener los correos de la carpeta" + folder.name)
        return emails
    

def classify_email(correo, date):
    global df

    try:
        if isinstance(correo, FileAttachment) or any(entorno in correo.subject for entorno in entornos["pre"]):
            return df

        subject = correo.subject
        body = correo.text_body

        resultado = "OK"
        urgente = "NO"
        observaciones = ""

        if correo.categories:
            if "KO" in correo.categories:
                resultado = "KO"
                error_match = re.search(r'(ERROR:|error:|Error:)(.*)', subject, re.IGNORECASE)
                observaciones = error_match.group(2) if error_match else "No hay mensaje de error, revisar el correo"
                # Eliminar del subject el mensaje de error
                subject = subject.split("Error:")[0] if "Error:" in subject else subject
                subject = subject.split("ERROR:")[0] if "ERROR:" in subject else subject
                subject = subject.split("error:")[0] if "error:" in subject else subject

        if "URGENT" in subject:
            urgente = "SI"

        tecnologia = next((key for key, values in categories.items() if any(value in subject for value in values)), "NO CLASIFICADO")

        if "Petició desplegament Genèric SIA a PRO" in subject:
            tecnologia = next((key for key, values in categories.items() if any(value in body for value in values)), "NO CLASIFICADO")

        check = "NO"
        incidencia = ""

        #Buscar si hay una incidencia resuelta
        incidencia_match = re.search(r'Aquesta petició resol una incidència\?[\s\r\n]*Sí', body, re.IGNORECASE)
        if incidencia_match:
            check = "SI"
            incidencia_match = re.search(r'Indiqueu el codi de la incidència: ([^\r\n]+)', body)
            incidencia = incidencia_match.group(1) if incidencia_match else "No hi ha incidència associada"

        # Buscar todas las incidencias numeradas
        incidencia_match_resta = re.findall(r'Nº incidència: ([^\r\n]+)|Nº incidencia: ([^\r\n]+)', body)
        if incidencia_match_resta:
            incidencias = [match[0] if match[0] else match[1] for match in incidencia_match_resta]
            incidencia = ", ".join(incidencias) if incidencias else "No hi ha incidència associada"
            if incidencia == "":
                incidencia = "Incidències múltiples"
            check = "SI"


        if "Fi Distribució tècnica paquet" in correo.subject:
            try:
                date = correo.datetime_received.strftime("%Y-%m-d")
            except AttributeError:
                date = correo.SentOn.strftime("%Y-%m-d")
        if date:
            match = re.search(r'\b([A-Z]\d+)\b', str(subject))
            codi_value = match.group(1) if match else "No hi ha codi"

            df = df.append({
                "Codi": codi_value,
                "Aplicació": subject,
                "Tecnologia": tecnologia,
                "Resultat": resultado,
                "Urgent": urgente,
                "Té incidència": check,
                "Incidencia associada": incidencia,
                "Observacions": observaciones,
                "Data": date
            }, ignore_index=True)

            
            if tecnologia == "Devops" or tecnologia == ".NET" or tecnologia == "Websphere":
                if "[Instalables+Scripts+Normal]" in subject:
                    df.loc[len(df)] = ["PRO", subject, "BBDD", resultado, urgente, check, incidencia, observaciones, date]
        return df
    except:
        sg.Popup("Error en la classificació dels correus")
        return df


def agregar_a_plantilla(df,fechaFinal,nomArchivo,hoja):
    try:
        df = df[df['Total Producció'] != 0]

        #cargamos el archivo
        try:
            wb = load_workbook(nomArchivo)
            ws = wb[hoja]
        except:
            sg.PopupOKCancel("L'arxiu Plantilla.xlsx no es troba en la mateixa carpeta que el programa, si us plau mogui l'arxiu a la mateixa carpeta que el programa i torni a executar-lo o premi OK per a continuar sense agregar a la plantilla",title="Error")
            if popup_result == "OK":
                sg.Popup("ADVERTIMENT: No s'agregarà a la hoja +"+hoja+" de la plantilla, pero si es generarà l'informe.")
                pass
            else:
                sys.exit()

        check = False

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if isinstance(row[0].value, datetime.datetime):
                row[0].value = row[0].value.strftime("%d/%m/%Y")
            else:
                row[0].value = row[0].value

            if datetime.datetime.strptime(fechaFinal.strftime("%d/%m/%Y"), "%d/%m/%Y") <= datetime.datetime.strptime(row[0].value, "%d/%m/%Y"):
                check = True

        if check:
            popup_result = sg.PopupOKCancel(
                "ATENCIÓ: Ja existeix un informe amb les dates seleccionades "+str(fechaFinal.strftime("%d/%m/%Y"))+".Es generarà un nou informe, però les dades a la plantilla d'Excel no s'actualitzaran !!. Les dates a introduir a la plantilla d'Excel només poden ser posteriors a la data de l'últim informe generat.",
                title="Advertència"
                )
            if popup_result == "OK":
                pass
            else:
                sys.exit()


        mesActual = datetime.datetime.now().month
        anoActual = datetime.datetime.now().year

        #si la hoja es Master, no eliminamos nada
        if hoja == "Master":
            pass
        else:
            for row in reversed(list(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column))):
                if row[0].value:
                    fecha = row[0].value.split("/")
                    if len(fecha) == 3:
                        year = int(fecha[2])
                        month = int(fecha[1])
                        if year < anoActual and  month == mesActual:
                            ws.delete_rows(row[0].row, 1)


        if check == False:
            #agregamos el dataframe al excel, y en la columna 1 ponemos la fecha final
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append([fechaFinal.strftime("%d/%m/%Y")] + r)

        #guardamos el excel
        wb.save(nomArchivo)
    except:
        sg.Popup("Error en agregar a la plantilla, asseguri's que l'arxiu Plantilla.xlsx aquesta tancat")


def dar_formato_excel(nombreArchivo, hoja="Tecnologies"):
    try:
        # Leer el archivo
        wb = load_workbook(nombreArchivo)
        ws = wb[hoja]

        # Quitar formato anterior a todas las filas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        # Seleccionar solo el máximo de columnas
        for cell in ws['1:1']:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

        # Dar formato a las filas de manera intercalada
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color='e6f2ff', end_color='e6f2ff', fill_type='solid')

        # En caso de que la primera columna sea "Total", darle un formato diferente
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[0].value == "Total":
                cell.fill = PatternFill(start_color='d9ead3', end_color='d9ead3', fill_type='solid')
                cell.font = Font(color='006100')


        # Guardar el archivo Excel con los estilos aplicados
        wb.save(nombreArchivo)
    except:
        sg.Popup("Error en donar format a l'Excel, asseguri's que l'arxiu aquesta tancat")



def set_style_to_excel(df, fechaInicio, fechaFinal):
     # crear carpeta si no existe
    try:
        if not os.path.exists("Informes_Generats"):
            os.makedirs("Informes_Generats")

        #ubicar el archivo en la carpeta
        os.chdir("Informes_Generats")


        #contar el total de urgents, es decir aquellos que Urgent = SI
        total_urgents = df[df["Urgent"] == "SI"].shape[0]

        #contar incidencias asociadas, es decir aquellos que Té incidència associada? = SI
        total_incidencies = df[df["Té incidència"] == "SI"].shape[0]

        #total desplegaments
        total_desplegaments = df.shape[0]

        #convertir la columna de Data a el formato de fecha d/m/Y
        df['Data'] = pd.to_datetime(df['Data']).dt.strftime('%d/%m/%Y')
        #Vamos a darle estilo al excel
        nomArchivo = "Informes_Gestió_Desplegament_"+fechaInicio.strftime("%d-%m-%Y")+"_"+fechaFinal.strftime("%d-%m-%Y")+".xlsx"
        writer = pd.ExcelWriter(nomArchivo, engine='xlsxwriter')
        df.to_excel(writer, sheet_name="Registres_"+fechaInicio.strftime("%d-%m-%Y")+"_"+fechaFinal.strftime("%d-%m-%Y"), index=False)
        workbook = writer.book
        worksheet = writer.sheets["Registres_" + fechaInicio.strftime("%d-%m-%Y") + "_" + fechaFinal.strftime("%d-%m-%Y")]
        # Add a header format.

        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#003f99',
            'border': 1,
             'font_color': '#ffffff'})

        #dar espacio a las columnas
        worksheet.set_column('A:A', 10)
        worksheet.set_column('B:B', 100)
        worksheet.set_column('C:C', 20)
        worksheet.set_column('D:D', 10)
        worksheet.set_column('E:E', 10)
        worksheet.set_column('F:F', 20)
        worksheet.set_column('G:G', 20)
        worksheet.set_column('H:H', 70)
        worksheet.set_column('I:I', 20)

        #Ponerle colores a las filas color='b3d2ff' , de manera intercalada
        for row_num in range(1, len(df)):
            if row_num % 2 == 0:
                worksheet.set_row(row_num, None, workbook.add_format({'bg_color': '#e6f2ff'}))
            else:
                worksheet.set_row(row_num, None, workbook.add_format({'bg_color': '#ffffff'}))

        # Write the column headers with the defined format.
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        writer.save()

        # abrir el excel y agregar una nueva hoja
        wb = Workbook()
        wb = load_workbook(nomArchivo)
        ws2 = wb.create_sheet("Resum")
        ws2['A1'] = "Tecnologia"
        ws2['B1'] = "Producció OK"
        ws2['C1'] = "Producció KO"
        ws2['D1'] = "Total Producció"
        ws2['E1'] = "Urgents"

        # Aplicar estilo a la cabecera
        for cell in ws2['1:1']:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')
            cell.font = Font(color='FFFFFF')

        # Ajustar el ancho de las columnas
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 20
        ws2.column_dimensions['D'].width = 20
        ws2.column_dimensions['E'].width = 20

         # Definir diccionario de tecnologías
        tecnologias = {
            "Devops": 2,
            "Paquet": 3,
            "Websphere": 4,
            "Client/Servidor": 5,
            "BIM": 6,
            ".NET": 7,
            "BBDD": 8,
            "Documentum": 9,
            "Cognos": 10,
            "Porlet": 11,
            "Pegats": 12,
            "Otros": 13,
            "BI": 14
        }

        # crear un nuevo dataset con [Tecnologia,Producció OK,Producció KO,Total Producció,Urgents] 
        dataResum = []
        for tecnologia, fila in tecnologias.items():
            tecnologia_ok = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "OK")].count()[0]
            tecnologia_ko = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "KO")].count()[0]
            tecnologia_urgent = df.loc[(df['Tecnologia'] == tecnologia) & (df['Urgent'] == "SI")].count()[0]
            dataResum.append([tecnologia, tecnologia_ok, tecnologia_ko, tecnologia_ok + tecnologia_ko, tecnologia_urgent])

        dfResum = pd.DataFrame(dataResum, columns=["Tecnologia", "Producció OK", "Producció KO", "Total Producció", "Urgent"])

        os.chdir("..")
        agregar_a_plantilla(dfResum,fechaFinal,"Plantilla.xlsx","Tecnologies")

        # Ordenar el dataset por Total Producció
        dfResum = dfResum.sort_values(by=['Total Producció'], ascending=False)

        # Eliminar valores con total 0
        dfResum = dfResum[dfResum['Total Producció'] != 0]

        # agregar una fila final con el total para cada columna
        dfResum.loc['Total'] = dfResum.sum()

        #poner nombre de Total en la ultima fila primera columna
        dfResum.iloc[-1, dfResum.columns.get_loc('Tecnologia')] = "Total"

        agregar_a_plantilla(dfResum,fechaFinal,"Plantilla.xlsx","Master")

        os.chdir("Informes_Generats")

        # ponemos el datraframe en el excel sin el encabezado
        for r in dataframe_to_rows(dfResum, index=False, header=False):
            ws2.append(r)

        # Aplicar estilo a la tabla
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color='e6f2ff', end_color='e6f2ff', fill_type='solid')


        # A partir del dataframe , crear un gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Resum"
        chart.y_axis.title = "Desplegaments"
        chart.x_axis.title = "Tecnología"

        # Agregar los datos y categorías al gráfico
        data = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=ws2.max_row)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Agregar el gráfico a la hoja
        ws2.add_chart(chart, "H2")

        ws2['A15'] = f"Desplegaments del {fechaInicio.strftime('%d-%m-%Y')} al {fechaFinal.strftime('%d-%m-%Y')}"

        if total_incidencies == 1:
            ws2['A17'] = f"De {total_desplegaments} desplegaments, {total_urgents} ha sigut urgent, d’aquests, {total_incidencies} tenia incidència associada."
        elif total_incidencies == 0:
            ws2['A17'] = f"De {total_desplegaments} desplegaments, {total_urgents} ha sigut urgent,d’aquests, cap tenia incidència associada."
        else:
            ws2['A17'] = f"De {total_desplegaments} desplegaments, {total_urgents} han sigut urgents, d’aquests, {total_incidencies} tenien incidència associada."

       

        # Guardar el excel
        wb.save(nomArchivo)

        #Vamos a crear una nueva hoja donde meteremos otra tabla
        ws3 = wb.create_sheet("Desplegament_KO")
        ws3['A1'] = "Petició"
        ws3['B1'] = "Aplicació"
        ws3['C1'] = "Tecnologia"
        ws3['D1'] = "Peticionari"
        ws3['E1'] = "Error"

        # Aplicar estilo a la cabecera
        for cell in ws3['1:1']:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')
            cell.font = Font(color='FFFFFF')

        # Ajustar el ancho de las columnas
        ws3.column_dimensions['A'].width = 20
        ws3.column_dimensions['B'].width = 60
        ws3.column_dimensions['C'].width = 30
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 50



        #Creamos un nuevo dataset con solo los KOs
        dfKO = df.loc[df['Resultat'] == "KO"]

        #ponermos los datos en el excel
        for r in dataframe_to_rows(dfKO, index=False, header=False):
            #agregamos solo las columnas que nos interesan
            if r[7] == "":
                r[7] = "No hay mensaje de error, revisar el correo"
            ws3.append([r[0],r[1],r[2],"",r[7]])

        # Aplicar estilo a la tabla
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color='e6f2ff', end_color='e6f2ff', fill_type='solid')

        # Guardar el excel
        wb.save(nomArchivo)
        os.chdir("..")

    except:
        sg.Popup("Error en generar l'Excel")


def get_date(val,fechaIncial=None):
    """Muestra un calendario para que el usuario seleccione una fecha."""
    root = Tk()
    root.geometry("300x350")
    if val == 0:
        root.title("Seleccionar data d'inici")
    else:
        root.title("Seleccionar data final")

    
    def get_selected_date():
        """Obtiene la fecha seleccionada por el usuario y cierra la ventana."""
        nonlocal selected_date
        selected_date = cal.selection_get()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        nonlocal selected_date
        selected_date = None
        root.destroy()
        exit()

     #Poner un titulo en la ventana
    if val == 0:
        title_label = Label(root, text="Seleccionar data d'inici", font=("Helvetica", 16), fg="black", bg="white")
        title_label.pack(pady=10)
        cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,maxdate=now)
        cal.pack(pady=20)
    else:
        title_label = Label(root, text="Seleccionar data final", font=("Helvetica", 16), fg="black", bg="white")
        title_label.pack(pady=10)

        #Solo se podra seleccionar una fecha igual o posterior a la fecha de inicio
        cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,mindate=fechaIncial,maxdate=now)
        cal.pack(pady=20)

    selected_date = None
    
    btn_ok = Button(root, text="OK", command=get_selected_date)
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="CANCEL·LAR", command=cancel)
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Heu de seleccionar una data"))

    root.mainloop()

    if selected_date is None:
        raise ValueError("No s'ha seleccionat cap data")
    return selected_date

def main():
    global df
    print("Iniciant programa...")
    print(".............................................")
    
    dominio = "corppro"
    correo = "gestioversions@bcn.cat"
    #Permitir que el usuario introduzca su nombre de usuario y contraseña 3 veces como máximo

    for i in range(3):
        print("ATENCIÓ : Numero d'intents restants: " + str(3-i))
        nombre_de_usuario = input("Introduce tu nombre de usuario Nominal : ")
        psswd = getpass("Introduce tu contrasena: ")
       
        try:
            credenciales = Credentials(username=f'{dominio}\\{nombre_de_usuario}', password=psswd )
            print("Connectant amb el servidor correu.bcn.cat ...")
            print(".............................................")
            config = Configuration(
                server="correu.bcn.cat",
                credentials=credenciales
            )

            cuenta = Account(
                primary_smtp_address=correo,
                config=config,
                autodiscover=False,
                access_type=DELEGATE
            )
            break
        except:
            print("Usuari o Contrasenya incorrecta")


    # Obtener la carpeta de entrada/ 
    luis_folder = cuenta.inbox / str(now.year) 


    print("Seleccione el rango de fechas")
    start_date = get_date(0)
    end_date = get_date(1,start_date)

    #convertir start_date y end_date a datetime
    start_date = parser.parse(str(start_date))
    end_date = parser.parse(str(end_date))

    print(".............................................")
    print("Dates seleccionades:")
    print(f"Data d'inici: {start_date}")
    print(f"Data final: {end_date}")

    start_time = time.time()
    try:
        print(f"Si us plau esperi, processant ....")
        print(".............................................")
        get_folder_emails_recursive(luis_folder, start_date, end_date)
    except ValueError:
        print("Ha ocorregut un error, si us plau intenti de nou o posi's en contacte amb el desenvolupador")
        return


    set_style_to_excel(df, start_date, end_date)
    
    print(".............................................")
    #pionerlo en minutos
    print(f"Tiempo de ejecución: {(time.time() - start_time)/60} minutos")
    

if __name__ == "__main__":
    main()
    dar_formato_excel("Plantilla.xlsx", "Tecnologies")
    dar_formato_excel("Plantilla.xlsx", "Master")
    input("Premeu qualsevol tecla per sortir...")

