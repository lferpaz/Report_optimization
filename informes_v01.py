import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,Alignment, Font
)
import win32com.client
from tkinter import Tk, Button, messagebox, Label, Entry, DISABLED
from tkcalendar import Calendar
import datetime
import os
import messagebox
import time
import warnings
warnings.filterwarnings("ignore")



now = datetime.datetime.now()

ENTORNO = "PRODUCCION"

# Global variables and information to the classificator
categories = {
    "Websphere": [".w61"],
    "NET": [".NET"],
    "ClientServidor": [".NT"],
    "BIM": [".BIM"],
    "Documentum": [".doc", ".wdk"],
    "Portlet": [".plr"],
    "Devops": [
        "Publicació d'API", "Petició desplegament DevOps",
        "DevOps", "Petició desplegament/execució scripts",
        "Petició desplegament/subscripció",
        "Petició desplegament/publicació d'API",
        "Petició de creació d'esquema BD"
    ],
    "Cognos": [".CBI", ".CDM"],
    "Paquet": [
        "Distribució", "Paquetització",
        "Crear petició de canvi per distribuir el paquet"
    ],
    "BBDD": [".BD", "Instalables+Scripts+Normal"]
}

entornos = {
    "pre": [
        "Munteu la maqueta", "Homologar", "Muntar la maqueta",
        "Muntar maqueta", "Assignació d'Assistència"
    ]
}

# Define dataframes containing the netx columns:
df = pd.DataFrame(
    columns=[
        "Entorno", "Aplicació", "Tecnologia", "Resultat", "Urgent",
        "Incidència associada?", "Observacions", "Data"
    ]
)


# Functions to read the files and create the dataframes
def connect_to_outlook():
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        return outlook
    except Exception as e:
        print("Error al conectar a Outlook:", e)
        return None


def get_inbox_folder(outlook, folder_name, subfolder_name=None):
    # in case of not finding the folder, it will return a mesaage
    try:
        inbox = outlook.Folders(folder_name).Folders(subfolder_name).Folders(str(now.year)) 
    except:
        print("La carpeta a la qual esteu intentant accedir no es troba, si us plau reviseu que el nom sigui el correcte.".format(folder_name, subfolder_name))
    return inbox


def get_all_messages_in_folder(folder, from_date, to_date):
    messages = []
    
    for item in folder.Items:
        if isinstance(item, win32com.client.CDispatch) and item.Class == 43:
            if item.SentOn.date() >= from_date and item.SentOn.date() <= to_date and not any(word in item.subject for word in entornos["pre"]):
                messages.append(item)      
    if folder.Folders.Count > 0:
        for subfolder in folder.Folders:
            messages.extend(get_all_messages_in_folder(subfolder, from_date, to_date))
    return messages


def get_inbox_messages(inbox, from_date, to_date):
    messages = []
    for folder in inbox.Folders:
        messages.extend(get_all_messages_in_folder(folder, from_date, to_date))
    return messages


def extract_emails(messages, start_date, end_date):
    emails_no_classified = []
    datas = []
    emails = []
    
    for message in messages:
        date_str = ""
        if "Horari seleccionat" in message.Body:
            try:
                date_str = message.Body.split("Horari seleccionat")[-1].split("\t")[1].split(" ")[0]
            except IndexError:
                date_str = message.Body.split("Horari seleccionat")[-1].split(" ")[1]
        elif "Es planifica el desplegament en l'horari:" in message.Body:
            date_str = message.Body.split("Es planifica el desplegament en l'horari:")[-1].split(" ")[1]
        elif "Per la data :" in message.Body:
            date_str = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]
            
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            try:
                date = datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                date = None
                
        if date and start_date <= date <= end_date:
            datas.append(date)
            emails.append(message)
            continue
        emails_no_classified.append(message)
        
    return emails, emails_no_classified, datas


def classify_message_and_inter_into_dataframe(messages, df):
    rows = []
    for message in messages:
        tecnologia = "NO DEFINIDO"
        observacions = message.Subject.split("Error:")[1].strip() if "Error:" in message.Subject else ""
        subject = message.Subject.split("Error:")[0].strip() or message.Subject.split("ERROR:")[0].strip()
        categories_list = message.Categories
        urgent = "SI" if subject.startswith("URGENT") else "NO"
        resultat=""
        
        if "KO" in categories_list:
            resultat = "KO"
        else:
            resultat = "OK"

        tecnologia = next((key for key in categories if any(word in subject for word in categories[key])), "NO DEFINIDO")
        
        if "Instalables+Scripts+Normal" in subject and tecnologia != "BBDD":
            data = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]
            data = datetime.datetime.strptime(data, "%d/%m/%Y").date()
            rows.append([ENTORNO, subject, "BBDD", resultat, urgent, "", observacions, data])
            
        rows.append([ENTORNO, subject, tecnologia, resultat, urgent, "", observacions, ""])
    
    new_df = pd.DataFrame(rows, columns=df.columns)
    
    return new_df

def pass_df_to_excel(df, from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Informes_Gestió_Versions_{from_date}_{to_date}"

    # Define the header of the Excel file
    ws['A1'] = "Entorno"
    ws['B1'] = "Aplicació"
    ws['C1'] = "Tecnologia"
    ws['D1'] = "Resultat"
    ws['E1'] = "Urgent"
    ws['F1'] = "Incidència associada?"
    ws['G1'] = "Observacions"
    ws['H1'] = "Data"

    # Iterate over the dataframe and append the data to the Excel file
    for index, row in df.iterrows():
        ws.append([
            row["Entorno"],
            row["Aplicació"],
            row["Tecnologia"],
            row["Resultat"],
            row["Urgent"],
            row["Incidència associada?"],
            row["Observacions"],
            row["Data"]
        ])

    # Define the font and alignment for the header
    font = Font(name='Arial', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='ffffff')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

    # Apply the style to the header
    for cell in ws['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')

    # Adjust the width of the columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 150
    ws.column_dimensions['H'].width = 20

    # Define the fill color for odd rows
    fill = PatternFill(start_color='b3d2ff', end_color='FFC000', fill_type='solid')

    # Apply the style to odd rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill

    # crear carpeta si no existe
    if not os.path.exists("Informes_Generados"):
        os.makedirs("Informes_Generados")
    
    wb.save(f"Informes_Generados/Informes_Gestió_Desplegament_{from_date}_{to_date}.xlsx")
    

def get_date(val):
    """Muestra un calendario para que el usuario seleccione una fecha."""
    root = Tk()
    root.geometry("300x300")
    if val == 0:
        root.title("Seleccionar data d'inici")
    else:
        root.title("Seleccionar data final")

    def get_selected_date():
        """Obtiene la fecha seleccionada por el usuario y cierra la ventana."""
        nonlocal selected_date
        selected_date = cal.selection_get()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        nonlocal selected_date
        selected_date = None
        root.destroy()
        exit()

    selected_date = None
    

    cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,maxdate=now)
    cal.pack(pady=20)

    btn_ok = Button(root, text="OK", command=get_selected_date)
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="CANCEL·LAR", command=cancel)
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Debes seleccionar una fecha"))

    root.mainloop()

    if selected_date is None:
        raise ValueError("No se ha seleccionado ninguna fecha")
    return selected_date



   




def main():
    global df
    # Mostramos mensajes de incializacion del programa
    print("###############################################")
    print("##                                           ##")
    print("##  Programa de generació d'informes de      ##")
    print("##  gestió de desplegaments d'aplicacions    ##")
    print("##                                           ##")
    print("###############################################")
    print()
     # Obtener la fecha seleccionada por el usuario y calcular la fecha inicial restando 4 días
    selected_date = get_date(0)
    print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")
 

    to_date = get_date(1)
    print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")

    #Espera 1 segon per a que es mostri la data final
    time.sleep(1)


    # Darle la opcion al usuario de revisar las fechas seleccionadas y poder cancelarlas y volver a seleccionarlas o simplemente continuar o terminar el programa
    # creamos un menu con las opciones de revisar las fechas, continuar o terminar el programa


    print("Opcions:")
    print("1. Revisar les dates")
    print("2. Continuar")
    print("3. Sortir")

    print()
    
    # creamos un bucle para que el usuario pueda elegir una opcion hasta que la opcion sea correcta
    while True:
        try:
            # pedimos al usuario que introduzca una opcion
            opcion = int(input("Introdueix una opció: "))
            # si la opcion es correcta salimos del bucle
            if opcion == 1:
                selected_date = get_date(0)
                print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")

                to_date = get_date(1)
                print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")
                time.sleep(1)
                continue
            elif opcion == 2:
                break
            elif opcion == 3:
                exit()
            else:
                print("Opció incorrecta")
                continue
        except ValueError:
            print("Opció incorrecta")
            continue
    

    
    from_date_week = selected_date - datetime.timedelta(days=7)
    
    # Conectar con Outlook y obtener la carpeta de la bandeja de entrada
    print("Connectant al vostre compte d'Outlook...")
    outlook = connect_to_outlook()
    inbox = get_inbox_folder(outlook, "gestioversions@bcn.cat","Bandeja de entrada")
   

    # Obtener los mensajes de la bandeja de entrada en el rango de fechas especificado
    print(f"Obtenint els missatges de {from_date_week.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    print()
    print("Si us plau espera un moment...")
    print()
    messages = get_inbox_messages(inbox, from_date_week, to_date)

    
    # Extraer información relevante de los correos electrónicos y clasificarlos en el dataframe
    print(f"Realitzar l'extracció dels mails de {selected_date.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    emails, emails_no_clasificados, datas = extract_emails(messages, selected_date, to_date)
    
    df = classify_message_and_inter_into_dataframe(emails, df)

    #Si el dataframe esta vacio, mostrar un mensaje de error y salir del programa
    if df.empty:
        messagebox.showerror("Error", "No s'han trobat correus electrònics al rang de dates especificat")
        exit()
    
    # Agregar la columna 'datas' en el orden de la lista, pero solo a aquellas filas del dataframe que no tengan fecha
    df.loc[df['Data'] == '', 'Data'] = datas
    
    # Pasar el dataframe a un archivo de Excel y guardar en disco
    pass_df_to_excel(df, selected_date, to_date)

    print("Fitxer Excel generat correctament !!")
    print()
    print("###############################################")
    print("##  fi del programa                          ##")
    print("###############################################")
    
    
    
#call the main function
if __name__ == "__main__":
    main()

    
 
        