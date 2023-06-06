import datetime
import os
import time
import warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side,NamedStyle

import win32com.client
from tkinter import Tk, Button, messagebox
from tkcalendar import Calendar
import babel
from babel.numbers import format_currency

#import win32timezone
import win32timezone
from openpyxl.chart import BarChart, Reference

warnings.filterwarnings("ignore")

now = datetime.datetime.now()

ENTORNO = "PRODUCCION"

# Global variables and information to the classificator
categories = {
    "Websphere": [".w61"],
    "NET": [".NET"],
    "ClientServidor": [".NT"],
    "BIM": [".BIM"],
    "Documentum": [".doc", ".wdk"],
    "Portlet": [".plr"],
    "Devops": [
        "Publicació d'API", "Petició desplegament DevOps",
        "DevOps", "Petició desplegament/execució scripts",
        "Petició desplegament/subscripció",
        "Petició desplegament/publicació d'API",
        "Petició de creació d'esquema BD",
        "Petició desplegament/creació esquema"
    ],
    "Cognos": [".CBI", ".CDM"],
    "Paquet": ["Fi Distribució tècnica paquet","Distribucio pegats seguretat anual"],
    "BBDD": [".BD", "Instalables+Scripts+Normal"]
}

entornos = {
    "pre": [
        "Munteu la maqueta", "Homologar", "Muntar la maqueta",
        "Muntar maqueta", "Assignació d'Assistència"
    ]
}

# Define dataframes containing the netx columns:
df = pd.DataFrame(
    columns=[
        "Entorn", "Aplicació", "Tecnologia", "Resultat", "Urgent",
        "Té incidència associada?","incidència associada", "Observacions", "Data"
    ]
)


# Functions to read the files and create the dataframes
def connect_to_outlook():
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        return outlook
    except Exception as e:
        print("Error al conectar a Outlook:", e)
        return None


def get_inbox_folder(outlook, folder_name, subfolder_name=None):
    # in case of not finding the folder, it will return a mesaage
    try:
        inbox = outlook.Folders(folder_name).Folders(subfolder_name).Folders(str(now.year)) 
    except:
        print("La carpeta a la qual esteu intentant accedir no es troba, si us plau reviseu que el nom sigui el correcte.".format(folder_name, subfolder_name))
    return inbox


def get_all_messages_in_folder(folder, from_date, to_date):
    messages = []
    
    for item in folder.Items:
        if isinstance(item, win32com.client.CDispatch) and item.Class == 43 and from_date <= item.SentOn.date() <= to_date:
            if  not any(word in item.subject for word in entornos["pre"]):
                messages.append(item)      
    if folder.Folders.Count > 0:
        for subfolder in folder.Folders:
            messages.extend(get_all_messages_in_folder(subfolder, from_date, to_date))
    return messages


def get_inbox_messages(inbox, from_date, to_date):
    messages = []
    for folder in inbox.Folders:
        messages.extend(get_all_messages_in_folder(folder, from_date, to_date))
    return messages


def extract_emails(messages, start_date, end_date):
    datas = []
    emails = []
    
    for message in messages:
        date_str = ""
        if "Horari seleccionat" in message.Body:
            try:
                date_str = message.Body.split("Horari seleccionat")[-1].split("\t")[1].split(" ")[0]
            except IndexError:
                date_str = message.Body.split("Horari seleccionat")[-1].split(" ")[1]
        elif "Es planifica el desplegament en l'horari:" in message.Body:
            date_str = message.Body.split("Es planifica el desplegament en l'horari:")[-1].split(" ")[1]
        elif "Per la data :" in message.Body:
            date_str = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]

        elif "Fi Distribució tècnica paquet" in message.Subject:
            #date_str sera la fecha que se envio el correo
            date_str = message.SentOn.strftime("%Y-%m-%d")

        elif "Distribucio pegats seguretat anual" in message.Subject:
            date_str = message.Body.split("Data i hora fi:")[-1].split(" ")[1].split(".")[0]
            
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            try:
                date = datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                date = None
                
        if date and start_date <= date <= end_date:
            datas.append(date)
            emails.append(message)
            continue
       
        
    return emails, datas


def classify_message_and_inter_into_dataframe(messages, df):
    rows = []
    for message in messages:
        tecnologia = "NO DEFINIT"
        observacions = message.Subject.split("Error:")[1].strip() if "Error:" in message.Subject else ""
        subject = message.Subject.split("Error:")[0].strip() or message.Subject.split("ERROR:")[0].strip()
        categories_list = message.Categories
        urgent = "SI" if subject.startswith("URGENT") else "NO"
        resultat=""
        incidencia = ""
        check = ""
        fecha=""

        
        if "KO" in categories_list:
            resultat = "KO"
        else:
            resultat = "OK"

        tecnologia = next((key for key in categories if any(word in subject for word in categories[key])), "NO DEFINIDO")
        
        if "Instalables+Scripts+Normal" in subject and tecnologia != "BBDD":
            data = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]
            data = datetime.datetime.strptime(data, "%d/%m/%Y").date()
            rows.append([ENTORNO, subject, "BBDD", resultat, urgent, "","", observacions, data])

        if tecnologia == "Devops":
            #buscamos en el mensaje si hay alguna incidencia asociada "Aquesta petició resol una incidència?"
            if "Aquesta petició resol una incidència?" in message.Body:
                #Tomamos la respuesta que es lo que esta entre "Aquesta petició resol una incidència?" y un espacio en blanco, ejemplo Aquesta petició resol una incidència? No
                check = message.Body.split("Aquesta petició resol una incidència?")[-1].split("\r")[0].strip(" ")
                #Si la respuesta es "SI" buscamos la incidencia que es lo que esta despues de "Indiqueu el codi de la incidència:"
                if check == "Sí":
                    #La incidencia es todo lo que esta despues de Indiqueu el codi de la incidència: y antes del primer \r
                    incidencia = message.Body.split("Indiqueu el codi de la incidència:")[-1].split("\r")[0]
                    #si la incidencia esta vacia o no tiene ningun valor ponemos "No hi ha incidència associada"
                    if not incidencia:
                        incidencia = "No hi ha incidència associada"
                        check = "No"
                    elif not any(char.isdigit() for char in incidencia) or len(incidencia) < 2:
                        incidencia = "No hi ha incidència associada"
                        check = "No"

                else:
                    incidencia = "No hi ha incidència associada"

        if tecnologia == "Paquet":
            if "Petició:" in message.body:
                subject = message.body.split("Petició:")[-1].split("\r")[0].strip(" ")
            else:
                subject = message.body.split("Tipus CI i Nom: ")[-1].split("\r")[0].strip(" ")

        rows.append([ENTORNO, subject, tecnologia, resultat, urgent, check,incidencia, observacions, ""])
    
    new_df = pd.DataFrame(rows, columns=df.columns)

    
    
    return new_df



def pass_df_to_excel(df, from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Informes_Gestió_Versions_{from_date}_{to_date}"

    # Define the header of the Excel file
    ws['A1'] = "Entorn"
    ws['B1'] = "Aplicació"
    ws['C1'] = "Tecnologia"
    ws['D1'] = "Resultat"
    ws['E1'] = "Urgent"
    ws['F1'] = "Té incidència associada?"
    ws['G1'] = "incidència associada"
    ws['H1'] = "Observacions"
    ws['I1'] = "Data"

    # Iterate over the dataframe and append the data to the Excel file
    for index, row in df.iterrows():
        ws.append([
            row["Entorn"],
            row["Aplicació"],
            row["Tecnologia"],
            row["Resultat"],
            row["Urgent"],
            row["Té incidència associada?"],
            row["incidència associada"],
            row["Observacions"],
            row["Data"]
        ])

    # Define the font and alignment for the header
    font = Font(name='Arial', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='ffffff')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

    # Apply the style to the header
    for cell in ws['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')

    # Adjust the width of the columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 100
    ws.column_dimensions['I'].width = 20


    # Define the fill color for odd rows
    fill = PatternFill(start_color='b3d2ff', end_color='FFC000', fill_type='solid')

    # Apply the style to odd rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill

    # crear carpeta si no existe
    if not os.path.exists("Informes_Generats"):
        os.makedirs("Informes_Generats")


    # Crear tabla
    ws2 = wb.create_sheet("Resumen")
    ws2['A1'] = "Tecnologia"
    ws2['B1'] = "Producció OK"
    ws2['C1'] = "Producció KO"
    ws2['D1'] = "Total Producció"

    # Aplicar estilo a la cabecera
    for cell in ws2['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')

    # Ajustar el ancho de las columnas
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20

    # Definir diccionario de tecnologías
    tecnologias = {
        "Devops": 2,
        "Paquet": 3,
        "Websphere": 4,
        "ClientServidor": 5,
        "BIM": 6,
        "NET": 7,
        "BBDD": 8,
        "Documentum": 9,
        "Cognos": 10,
        "Porlet": 11,
    }

    # Iterar sobre las tecnologías en el diccionario
    for tecnologia, fila in tecnologias.items():
        # Ponemos la tecnología en la tabla
        ws2.cell(row=fila, column=1).value = tecnologia

        # Contar los OK y KO de la tecnología en el DataFrame
        tecnologia_ok = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "OK")].count()[0]
        tecnologia_ko = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "KO")].count()[0]

        if tecnologia_ok + tecnologia_ko == 0:
            ws2.delete_rows(fila)
            continue
        else:
            # Poner la suma en la columna OK, KO y Total
            ws2.cell(row=fila, column=2).value = tecnologia_ok
            ws2.cell(row=fila, column=3).value = tecnologia_ko
            ws2.cell(row=fila, column=4).value = tecnologia_ok + tecnologia_ko

    
    # Aplicar estilo a las filas impares
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill


    #eliminar las filas donde el "Total Producció" es 0
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        if row[3].value == 0:
            ws2.delete_rows(row[0].row)

    contador= 0
    #eliminar las filas que no tengan ningun valor
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        if row[0].value == None:
            ws2.delete_rows(row[0].row)
            contador = contador + 1


    # Obtener los datos de la hoja de cálculo y convertirlos en una lista de tuplas
    data = []
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row-contador, min_col=1, max_col=4):
        data.append((row[0].value, row[1].value, row[2].value, row[3].value))

    #Eliminar los none de la lista
    data = [x for x in data if x[0] is not None]

    # Ordenar la lista de tuplas por la columna "Total Producció" en orden descendente
    sorted_data = sorted(data, key=lambda x: x[3], reverse=True)

    # Reescribir la hoja de cálculo con los datos ordenados
    for i, row in enumerate(sorted_data, start=2):
        ws2.cell(row=i, column=1).value = row[0]
        ws2.cell(row=i, column=2).value = row[1]
        ws2.cell(row=i, column=3).value = row[2]
        ws2.cell(row=i, column=4).value = row[3]

    #eliminar las filas repetidas
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        if row[0].value == row[1].value:
            ws2.delete_rows(row[0].row)
            


    # Agregar una ultima fila con el total de cada columna
    ws2.cell(row=ws2.max_row + 1, column=1).value = "TOTAL"
    ws2.cell(row=ws2.max_row, column=2).value = "=SUM(B2:B"+str(ws2.max_row-1)+")"
    ws2.cell(row=ws2.max_row, column=3).value = "=SUM(C2:C"+str(ws2.max_row-1)+")"
    ws2.cell(row=ws2.max_row, column=4).value = "=SUM(D2:D"+str(ws2.max_row-1)+")"

    # A partir de la tabla generada en la hoja u "Resumen", crear un gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Resum"
    chart.y_axis.title = "Número d'Incidències"
    chart.x_axis.title = "Tecnología"

    # Agregar los datos y categorías al gráfico
    data = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws2.add_chart(chart, "G2")

    wb.save(f"Informes_Generats/Informes_Gestió_Desplegament_{from_date}_{to_date}.xlsx")
    

def get_date(val):
    """Muestra un calendario para que el usuario seleccione una fecha."""
    root = Tk()
    root.geometry("300x300")
    if val == 0:
        root.title("Seleccionar data d'inici")
    else:
        root.title("Seleccionar data final")

    def get_selected_date():
        """Obtiene la fecha seleccionada por el usuario y cierra la ventana."""
        nonlocal selected_date
        selected_date = cal.selection_get()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        nonlocal selected_date
        selected_date = None
        root.destroy()
        exit()

    selected_date = None
    

    cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,maxdate=now)
    cal.pack(pady=20)

    btn_ok = Button(root, text="OK", command=get_selected_date)
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="CANCEL·LAR", command=cancel)
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Heu de seleccionar una data"))

    root.mainloop()

    if selected_date is None:
        raise ValueError("No s'ha seleccionat cap data")
    return selected_date




def main():
    global df
    # Mostramos mensajes de incializacion del programa
    print("###############################################")
    print("##                                           ##")
    print("##  Programa de generació d'informes de      ##")
    print("##  gestió de desplegaments d'aplicacions    ##")
    print("##                                           ##")
    print("###############################################")
    print()
     # Obtener la fecha seleccionada por el usuario y calcular la fecha inicial restando 4 días
    selected_date = get_date(0)
    print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")

    to_date = get_date(1)
    print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")

    #Espera 1 segon per a que es mostri la data final
    time.sleep(1)

    # Darle la opcion al usuario de revisar las fechas seleccionadas y poder cancelarlas y volver a seleccionarlas o simplemente continuar o terminar el programa
    # creamos un menu con las opciones de revisar las fechas, continuar o terminar el programa
    print("Opcions:")
    print("1. Revisar les dates")
    print("2. Continuar")
    print("3. Sortir")
    print()
    # creamos un bucle para que el usuario pueda elegir una opcion hasta que la opcion sea correcta
    while True:
        try:
            # pedimos al usuario que introduzca una opcion
            opcion = int(input("Introdueix una opció: "))
            # si la opcion es correcta salimos del bucle
            if opcion == 1:
                selected_date = get_date(0)
                print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")

                to_date = get_date(1)
                print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")
                time.sleep(1)
                continue
            elif opcion == 2:
                break
            elif opcion == 3:
                exit()
            else:
                print("Opció incorrecta")
                continue
        except ValueError:
            print("Opció incorrecta")
            continue
    
    from_date_week = selected_date - datetime.timedelta(days=7)
    
    # Conectar con Outlook y obtener la carpeta de la bandeja de entrada
    print("Connectant al vostre compte d'Outlook...")
    outlook = connect_to_outlook()
    inbox = get_inbox_folder(outlook, "gestioversions@bcn.cat","Bandeja de entrada")
   

    # Obtener los mensajes de la bandeja de entrada en el rango de fechas especificado
    print(f"Obtenint els missatges de {from_date_week.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    print()
    print("Si us plau espera un moment...")
    print()
    messages = get_inbox_messages(inbox, from_date_week, to_date)

    # Extraer información relevante de los correos electrónicos y clasificarlos en el dataframe
    print(f"Realitzar l'extracció dels mails de {selected_date.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    emails,datas = extract_emails(messages, selected_date, to_date)

    df = classify_message_and_inter_into_dataframe(emails, df)



    print("Extracció realitzada correctament, se ha creat al dataset !!")

    #Si el dataframe esta vacio, mostrar un mensaje de error y salir del programa
    if df.empty:
        messagebox.showerror("Error", "No s'han trobat correus electrònics al rang de dates especificat")
        exit()
    
    # Agregar la columna 'datas' en el orden de la lista, pero solo a aquellas filas del dataframe que no tengan fecha
    df.loc[df['Data'] == '', 'Data'] = datas
    
    # Pasar el dataframe a un archivo de Excel y guardar en disco
    pass_df_to_excel(df, selected_date, to_date)

    print("Fitxer Excel generat correctament !!")
    print()
    print("###############################################")
    print("##  fi del programa                          ##")
    print("###############################################")
    
    
if __name__ == "__main__":
    # Ejecutar el programa y hacer un try/except para capturar cualquier excepción
    try:
        main()
        input("Premeu una tecla per a continuar...")
    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un error: {e}")
        exit()
