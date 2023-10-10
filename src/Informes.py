import datetime
import locale
import os
import pandas as pd
import PySimpleGUI as sg
from babel import numbers
from datetime import timedelta
from openpyxl import Workbook,load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import time
import warnings
import win32com.client
import win32timezone
from tkinter import Tk, Button, messagebox, Label
from tkcalendar import Calendar, DateEntry
from dateutil import parser


warnings.filterwarnings("ignore")

# GLOBAL VARIABLES
now = datetime.datetime.now()
selected_date = None
to_date = None
generar_desplegaments_PRO = False

ENTORNO = "PRODUCCION"


# DATA STRUCTURES
categories = {
    "Websphere": [".w61"],
    ".NET": [".NET"],
    "Client/Servidor": [".NT"],         
    "BIM": [".BIM"],                  
    "Documentum": [".doc", ".wdk"],
    "Portlet": [".plr"],                
    "Devops": [
        "Publicació d'API", "Petició desplegament DevOps",
        "DevOps", "Petició desplegament/execució scripts",
        "Petició desplegament/subscripció",
        "Petició desplegament/publicació d'API",
        "Petició de creació d'esquema BD",
        "Petició desplegament/creació esquema"
    ],
    "Cognos": [".CBI", ".CDM"],    
    "Paquet": ["Fi Distribució tècnica paquet"],
    "BBDD": [".BD", "Instalables+Scripts+Normal"],
    "Pegats" : ["Distribucio pegats seguretat"],
    "Otros" : ["SILTRA","TREM","FICOR"]
}


entornos = {
    "pre": [
        "Munteu la maqueta", "Homologar", "Muntar la maqueta",
        "Muntar maqueta", "Assignació d'Assistència","Petició desplegament genèric SIA a PRE","Detindre aplicació","RE:","Ok detenció aplicació a PRO","Desinstal·lació aplicació"
    ]
}

df = pd.DataFrame(
    columns=[
        "Entorn", "Aplicació", "Tecnologia", "Resultat", "Urgent",
        "Té incidència associada?","incidència associada", "Observacions", "Data"
    ]
)

df_despliegues = pd.DataFrame(
    columns=[
        "Responsable","Petició","Tecnologia","Reinici de servidor","Publicació d'API","Subscripció d'API","GetAccess","API + Versio"
    ]
)


# FUNCTIONS TO GET THE MESSAGES FROM OUTLOOK
def connect_to_outlook():
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        return outlook
    except Exception as e:
        print("Error al conectar a Outlook:", e)
        return None


def get_inbox_folder(outlook, folder_name, subfolder_name=None):
    inbox = None
    try:
        inbox = outlook.Folders(folder_name).Folders(subfolder_name).Folders(str(now.year)) 
    except:
        print("La carpeta a la qual esteu intentant accedir no es troba, si us plau reviseu que el nom sigui el correcte.".format(folder_name, subfolder_name))
    return inbox


def get_all_messages_in_folder(folder, from_date, to_date):
    messages = []
    if folder is not None:
        items = folder.Items
        if items is not None:
            for item in items:
                if isinstance(item, win32com.client.CDispatch) and item.Class == 43:
                    if not any(word in item.subject for word in entornos["pre"]):
                        messages.append(item)      
            if folder.Folders.Count > 0:
                for subfolder in folder.Folders:
                    messages.extend(get_all_messages_in_folder(subfolder, from_date, to_date))
    return messages


def get_inbox_messages(inbox, from_date, to_date):
    messages = []
    if not isinstance(inbox, win32com.client.CDispatch):
        print("Error: la carpeta de entrada no es válida")
        return messages
    
    for folder in inbox.Folders:
        messages.extend(get_all_messages_in_folder(folder, from_date, to_date))
    return messages



# FUNCTIONS TO EXTRACT THE DATE OF THE MESSAGES
def extract_emails(messages, start_date, end_date):
    emails_no_classified = []
    datas = []
    emails = []

    for message in messages:
        date_str = ""
        if "Horari seleccionat" in message.Body:
            try:
                date_str = message.Body.split("Horari seleccionat")[-1].split("\t")[1].split(" ")[0]
            except IndexError:
                date_str = message.Body.split("Horari seleccionat")[-1].split(" ")[1]
        elif "Es planifica el desplegament en l'horari:" in message.Body:
            date_str = message.Body.split("Es planifica el desplegament en l'horari:")[-1].split(" ")[1]
        elif "Per la data :" in message.Body:
            date_str = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]

        elif "Fi Distribució tècnica paquet"  in message.Subject:
            #date_str sera la fecha que se envio el correo
            date_str = message.SentOn.strftime("%Y-%m-%d")
        elif "Petició desplegament genèric" in message.Subject:
            #date_str sera la fecha que se envio el correo
            date_str = message.SentOn.strftime("%Y-%m-%d")
        elif "Distribucio pegats seguretat anual" in message.Subject:
            date_str = message.Body.split("Data i hora fi:")[-1].split(" ")[1].split(".")[0]
 
        try:
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            try:
                date = datetime.datetime.strptime(date_str, "%d/%m/%Y").date()
            except ValueError:
                date = None
                
        # Si la fecha es despues o igual que la fecha de inicio y antes o igual que la fecha de fin
        if date is not None and date >= start_date and date <= end_date:
            datas.append(date)
            emails.append(message)
            continue
        emails_no_classified.append(message)
        
    return emails, datas


# FUNCTIONS TO CLASSIFY THE MESSAGES BY TECHNOLOGY
def classify_message_and_inter_into_dataframe(messages, df):
    #################
    # INFORMES #
    rows = []
    for message in messages:
        tecnologia = "NO DEFINIT"
        observacions = message.Subject.split("Error:")[1].strip() if "Error:" in message.Subject else ""
        subject = message.Subject.split("Error:")[0].strip() or message.Subject.split("ERROR:")[0].strip()
        categories_list = message.Categories
        urgent = "SI" if subject.startswith("URGENT") else "NO"
        resultat=""
        incidencia = ""
        check = ""

        if "KO" in categories_list:
            resultat = "KO"
        else:
            resultat = "OK"
            
        tecnologia = next((key for key in categories if any(word in subject for word in categories[key])), "NO DEFINIDO")

        if "Petició desplegament Genèric SIA a PRO" in subject:
            #Buscamos un .NT en el cuerpo del mensaje
            if ".NT" in message.Body:
                tecnologia = "Client/Servidor"
            #Buscamos un .BIM en el cuerpo del mensaje
            elif ".BIM" in message.Body:
                tecnologia = "BIM"
            elif ".plr" in message.Body:
                tecnologia = "Porlet"
            elif ".CBI" or ".CDM" in message.Body:
                tecnologia = "Cognos"
            
        
        if "Instalables+Scripts+Normal" in subject and tecnologia != "BBDD":
            data = message.Body.split("Per la data :")[-1].split(" ")[1].split(".")[0]
            data = datetime.datetime.strptime(data, "%d/%m/%Y").date()
            rows.append([ENTORNO, subject, "BBDD", resultat, urgent, "","", observacions, data])

        #Si solo pone "Scripts+Normal" en el asunto entonces la tecnologia es BBDD
        if "[Scripts+Normal]" in subject: 
            tecnologia = "BBDD"

        
            
        if "Aquesta petició resol una incidència?" in message.Body:
            
            check = message.Body.split("Aquesta petició resol una incidència?")[-1].split("\r")[0].strip(" ")
            if check == "Sí":
                incidencia = message.Body.split("Indiqueu el codi de la incidència:")[-1].split("\r")[0]
                
                if not incidencia:
                    incidencia = "No hi ha incidència associada"
                    check = "No"
                elif not any(char.isdigit() for char in incidencia) or len(incidencia) < 2:
                    incidencia = "No hi ha incidència associada"
                    check = "No"
            else:
                incidencia = "No hi ha incidència associada"

        if tecnologia == "Paquet":
            if "Petició:" in message.body:
                subject = message.body.split("Petició:")[-1].split("\r")[0].strip(" ")


        rows.append([ENTORNO, subject, tecnologia, resultat, urgent, check,incidencia, observacions, ""])
    
        
        #################
        # DESPLEGAMENTS #
        # Crear un diccionario con las categorías y sus valores por defecto
        categorias_valores = {
            "Desplegament amb reinici de servidors": "NO",
            "Subscripció d'API": "NO",
            "API": "NO",
            "GetAccess": "NO",
            "API + VERSIÓN": "NO"
        }
        
       
        categories_list = categories_list.split(";")

        #eliminar los espacios en blanco de la lista
        categories_list = [x.strip(' ') for x in categories_list]

        for categoria in categories_list:
            if categoria in categorias_valores:
                categorias_valores[categoria] = "SI"
                #eliminar la categoria de la lista
                categories_list.remove(categoria)
        
        # Obtener los valores de las categorías
        reinicio = categorias_valores["Desplegament amb reinici de servidors"]
        subscripcio = categorias_valores["Subscripció d'API"]
        api = categorias_valores["API"]
        getaccess = categorias_valores["GetAccess"]
        version = categorias_valores["API + VERSIÓN"]

        if getaccess == "SI":
            #buscamos en el cuerpo del mensaje la palabra "Acció J2EE:" y cogemos el valor que hay despues
            accio = message.Body.split("Acció J2EE:")[-1].split("\r")[0].strip(" ")
            getaccess = "SI - " + accio


        if "KO" in categories_list:
            categories_list.remove("KO")

        responsable = categories_list

        responsable = str(responsable).strip('[]')
        # Eliminar comillas simples y dobles de la cadena
        responsable = responsable.replace("'", "").replace('"', '')
        
        # Agregar los valores al DataFrame
        df_despliegues.loc[len(df_despliegues)] = [responsable,subject,tecnologia , reinicio, api, subscripcio, getaccess, version]

        #################
    
    new_df = pd.DataFrame(rows, columns=df.columns)
    return new_df , df_despliegues



def pass_df_to_excel(df, from_date, to_date):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Informes_Gestió_Versions_{from_date}_{to_date}"

    #contar el total de urgents, es decir aquellos que Urgent = SI
    total_urgents = df[df["Urgent"] == "SI"].shape[0]

    #contar incidencias asociadas, es decir aquellos que Té incidència associada? = SI
    total_incidencies = df[df["Té incidència associada?"] == "Sí"].shape[0]

    #total desplegaments
    total_desplegaments = df.shape[0]

    # Define the header of the Excel file
    ws['A1'] = "Entorn"
    ws['B1'] = "Aplicació"
    ws['C1'] = "Tecnologia"
    ws['D1'] = "Resultat"
    ws['E1'] = "Urgent"
    ws['F1'] = "Té incidència associada?"
    ws['G1'] = "incidència associada"
    ws['H1'] = "Observacions"
    ws['I1'] = "Data"

    # Iterate over the dataframe and append the data to the Excel file
    for index, row in df.iterrows():
        ws.append([
            row["Entorn"],
            row["Aplicació"],
            row["Tecnologia"],
            row["Resultat"],
            row["Urgent"],
            row["Té incidència associada?"],
            row["incidència associada"],
            row["Observacions"],
            row["Data"]
        ])

    # Define the font and alignment for the header
    font = Font(name='Arial', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='ffffff')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

    # Apply the style to the header
    for cell in ws['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')

    # Adjust the width of the columns
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 100
    ws.column_dimensions['I'].width = 20


    # Define the fill color for odd rows
    fill = PatternFill(start_color='b3d2ff', end_color='FFC000', fill_type='solid')

    # Apply the style to odd rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill

    # crear carpeta si no existe
    if not os.path.exists("Informes_Generats"):
        os.makedirs("Informes_Generats")


    # Crear tabla
    ws2 = wb.create_sheet("Resumen")
    ws2['A1'] = "Tecnologia"
    ws2['B1'] = "Producció OK"
    ws2['C1'] = "Producció KO"
    ws2['D1'] = "Total Producció"
    ws2['E1'] = "Urgents"

    # Aplicar estilo a la cabecera
    for cell in ws2['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='003f99', end_color='0000ff', fill_type='solid')

    # Ajustar el ancho de las columnas
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20

    # Definir diccionario de tecnologías
    tecnologias = {
        "Devops": 2,
        "Paquet": 3,
        "Websphere": 4,
        "Client/Servidor": 5,
        "BIM": 6,
        ".NET": 7,
        "BBDD": 8,
        "Documentum": 9,
        "Cognos": 10,
        "Porlet": 11,
        "Pegats": 12,
        "Otros": 13
    }

    # crear un nuevo dataset con [Tecnologia,Producció OK,Producció KO,Total Producció,Urgents] 
    dataResum = []
    for tecnologia, fila in tecnologias.items():
        tecnologia_ok = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "OK")].count()[0]
        tecnologia_ko = df.loc[(df['Tecnologia'] == tecnologia) & (df['Resultat'] == "KO")].count()[0]
        tecnologia_urgent = df.loc[(df['Tecnologia'] == tecnologia) & (df['Urgent'] == "SI")].count()[0]
        dataResum.append([tecnologia, tecnologia_ok, tecnologia_ko, tecnologia_ok + tecnologia_ko, tecnologia_urgent])

    dfResum = pd.DataFrame(dataResum, columns=["Tecnologia", "Producció OK", "Producció KO", "Total Producció", "Urgent"])

    # Ordenar el dataset por Total Producció
    dfResum = dfResum.sort_values(by=['Total Producció'], ascending=False)

    # Eliminar valores con total 0
    dfResum = dfResum[dfResum['Total Producció'] != 0]


    ''' 
    Agregar datos a el documento Plantilla de Excel
    '''

    # Cargar el archivo Excel existente
    archivo_excel = "Plantilla.xlsx"
    try:
        wb2 = load_workbook(archivo_excel)
        ws3 = wb2['Tecnologies']
        existe_registro = False

    except:
        sg.PopupOKCancel(
            "No existeix el fitxer Plantilla.xlsx a la carpeta Informes_Generats, si us plau, creeu-lo a partir de la plantilla que trobareu a la carpeta Plantilla.",
            title="Error"
        )
        sys.exit()


    to_date_ = to_date.strftime("%Y-%m-%d")
    from_date_ = from_date.strftime("%Y-%m-%d")

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=1):
        # convertimos string a objeto datetime
        valor_fecha = str(row[0].value)
        try:
            r_data = datetime.datetime.strptime(valor_fecha, '%Y-%m-%d %H:%M:%S').date()
            # Poner en formato dd/mm/yyyy
            r_data = r_data.strftime("%Y-%m-%d")
        except:
            r_data = datetime.datetime.strptime(valor_fecha, '%d/%m/%Y').date()
            # Poner en formato dd/mm/yyyy
            r_data = r_data.strftime("%Y-%m-%d")

        if r_data >= to_date_ or r_data >= from_date_:
            # Mostrar mensaje de error
            popup_result = sg.PopupOKCancel(
            "ATENCIÓ: Ja existeix un informe amb les dates seleccionades "+str(r_data)+".Es generarà un nou informe, però les dades a la plantilla d'Excel no s'actualitzaran !!. Les dates a introduir a la plantilla d'Excel només poden ser posteriors a la data de l'últim informe generat.",
            title="Advertència"
            )
            if popup_result == "OK":
                existe_registro = True
                break
            else:
                sys.exit()
        
    if not existe_registro:
        fechaActual = to_date.strftime("%d/%m/%Y")
        # Obtener la próxima fila vacía después de borrar las filas
        next_row = ws3.max_row + 1

        #Borrar los registros del mes del ano anterior, es decir si estamos en enero 2021, borrar los registros de enero 2020
        # Obtener la fecha actual
        today = datetime.date.today()
        mes_actual = today.month
        ano_actual = today.year

        filas_a_eliminar = []

        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=1):
            valor_fecha = row[0].value

            if valor_fecha is not None:
                if isinstance(valor_fecha, datetime.datetime):
                    fecha = valor_fecha.date()
                elif isinstance(valor_fecha, str):
                    try:
                        fecha = datetime.datetime.strptime(valor_fecha, '%Y-%m-%d').date()
                    except ValueError:
                        fecha = datetime.datetime.strptime(valor_fecha, '%d/%m/%Y').date()
                else:
                    continue  # Salta las filas con valores de fecha no válidos
                
                if fecha.month == mes_actual and fecha.year == (ano_actual - 1):
                    # Agrega el índice de la fila a la lista de filas a eliminar
                    filas_a_eliminar.append(row[0].row)

        # Elimina las filas en orden inverso para evitar problemas de índices cambiantes
        filas_a_eliminar.reverse()
        for idx in filas_a_eliminar:
            ws3.delete_rows(idx)

                
        # Agregar los nuevos datos al Excel
        for _, row_data in dfResum.iterrows():
            ws3.append([
                #cambiar formato de to?date de yyyy-mm-dd a dd/mm/yyyy pero manteniendo el tipo datetime 
                fechaActual,
                row_data["Tecnologia"],
                row_data["Producció OK"],
                row_data["Producció KO"],
                row_data["Total Producció"],
                row_data["Urgent"]
            ])

        # Aplicar estilo a los datos agregados
        for row in ws3.iter_rows(min_row=next_row, max_row=ws3.max_row, min_col=1, max_col=6):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = fill
        # Guardar el archivo Excel
        wb2.save(archivo_excel)
    

        #hacemos lo mismo para el excel en la hoja Master
        wb3 = load_workbook(archivo_excel)
        ws4 = wb3['Master']

        # Obtener la próxima fila vacía después de borrar las filas
        next_row = ws4.max_row + 1

        # Agregar los nuevos datos al Excel
        for _, row_data in dfResum.iterrows():
            ws4.append([
                fechaActual,
                row_data["Tecnologia"],
                row_data["Producció OK"],
                row_data["Producció KO"],
                row_data["Total Producció"],
                row_data["Urgent"]
            ])

        # Aplicar estilo a los datos agregados
        for row in ws4.iter_rows(min_row=next_row, max_row=ws4.max_row, min_col=1, max_col=6):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = fill

        # agregar una fila final con el total
        ws4.append([
            fechaActual,
            "Total",
            dfResum["Producció OK"].sum(),
            dfResum["Producció KO"].sum(),
            dfResum["Total Producció"].sum(),
            dfResum["Urgent"].sum()
        ])

        # Guardar el archivo Excel
        wb3.save(archivo_excel)

        

    # agregar una fila final con el total para cada columna
    dfResum.loc['Total'] = dfResum.sum()

    #poner nombre de Total en la ultima fila primera columna
    dfResum.iloc[-1, dfResum.columns.get_loc('Tecnologia')] = "Total"


    # ponemos el datraframe en el excel sin el encabezado
    for r in dataframe_to_rows(dfResum, index=False, header=False):
        ws2.append(r)
        
    # Aplicar estilo a la tabla
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill

    
    # A partir del dataframe , crear un gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Resum"
    chart.y_axis.title = "Desplegaments"
    chart.x_axis.title = "Tecnología"

    # Agregar los datos y categorías al gráfico
    data = Reference(ws2, min_col=2, min_row=1, max_col=4, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Agregar el gráfico a la hoja
    ws2.add_chart(chart, "H2")

    # concatenar la siguiente cadena Desplegaments del poner fecha en formato dd/mm/yyyy al poner fecha en formato dd/mm/yyyy
    from_date2 = df['Data'].min().strftime("%d/%m/%Y")
    to_date2 = df['Data'].max().strftime("%d/%m/%Y")
    
    ws2['A15'] = f"Desplegaments del {from_date2} al {to_date2}"

    # concatenar la siguiente cadena De n desplegaments, n han sigut urgents, d’aquests, n tenien incidència associada
    ws2['A17'] = f"De {total_desplegaments} desplegaments, {total_urgents} han sigut urgents, d’aquests, {total_incidencies} tenien incidència associada."

    wb.save(f"Informes_Generats/Informes_Gestió_Desplegament_{from_date}_{to_date}.xlsx")
    


def pass_df_despliegues_to_excel(df_despliegues, from_date, to_date):
    wb = Workbook()
    ws3 = wb.active
    ws3.title = f"Informes_desplegaments"

    # Define the font and alignment for the header
    font = Font(name='Arial', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='ffffff')
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    fill_2 = PatternFill(start_color='fae4b9', end_color='FFC000', fill_type='solid')

    # Apply the style to odd rows
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill_2

    # crear carpeta si no existe
    if not os.path.exists("Informes_Generats"):
        os.makedirs("Informes_Generats")

    # Definir la cabecera de la hoja
    ws3['A1'] = "Responsable"
    ws3['B1'] = "Petició"
    ws3['C1'] = "Tenologia"
    ws3['D1'] = "Reinici de servidor"
    ws3['E1'] = "Publicació d'API"
    ws3['F1'] = "Subscripció d'API"
    ws3['G1'] = "GetAccess"
    ws3['H1'] = "API + Versió"

    # Aplicar estilo a la cabecera
    for cell in ws3['1:1']:
        cell.font = font
        cell.alignment = alignment
        cell.fill = PatternFill(start_color='ffb521', end_color='0000ff', fill_type='solid')

    # Ajustar el ancho de las columnas
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 100
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 20
    ws3.column_dimensions['F'].width = 20
    ws3.column_dimensions['G'].width = 20
    ws3.column_dimensions['H'].width = 20

    # Agregar los datos al excel
    for index, row in df_despliegues.iterrows():
        ws3.append([
            row["Responsable"],
            row["Petició"],
            row["Tecnologia"],
            row["Reinici de servidor"],
            row["Publicació d'API"],
            row["Subscripció d'API"],
            row["GetAccess"],
            row["API + Versio"]
        ])

    # Aplicar estilo a la tabla
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill_2
                
    wb.save(f"Informes_Generats/Informes_desplegaments_PRO_{from_date}_{to_date}.xlsx")


def get_date_range(start_date=None, end_date=None):
    """Muestra un widget de rango de fechas para que el usuario seleccione un rango."""
    root = Tk()
    root.geometry("500x500")
    root.title("Gestió de desplegaments")

    def get_selected_dates():
        """Obtiene las fechas seleccionadas por el usuario y cierra la ventana."""
        nonlocal start_date, end_date
        start_date = start_entry.get_date()
        end_date = end_entry.get_date()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        # pregunta si se quiere salir sin seleccionar una fecha
        if messagebox.askokcancel("Sortir", "Vols sortir?"):
            nonlocal start_date, end_date
            start_date = None
            end_date = None
            root.destroy()
            sys.exit()

    now = datetime.datetime.now()
    if start_date is None:
        start_date = now - timedelta(days=4)
        end_date = now

    title_label = Label(root, text="Seleccionar rang de dates", font=("Helvetica", 16), fg="black", bg="white")
    title_label.pack(pady=10)

    start_label = Label(root, text="Data d'inici:", font=("Helvetica", 14 ), fg="green", bg="white")
    start_label.pack(pady=10)

    locale.setlocale(locale.LC_TIME, 'ca_ES')
    start_entry = DateEntry(root, width=12, background='white', foreground='black', date_pattern='dd/mm/yyyy', maxdate=now,locale='ca_ES')
    start_entry.set_date(start_date)
    start_entry.pack(pady=5)

    end_label = Label(root, text="Data final:",font=("Helvetica", 14 ), fg="green", bg="white")
    end_label.pack(pady=10)

    end_entry = DateEntry(root, width=12, background='white', foreground='black', date_pattern='dd/mm/yyyy', maxdate=now,locale='ca_ES')
    end_entry.set_date(end_date)
    end_entry.pack(pady=5)

    btn_ok = Button(root, text="Ok", command=get_selected_dates, width=5, height=1, bg="green", fg="white", font=("Helvetica", 10))
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="Cancel·la", command=cancel, width=10, height=1, bg="red", fg="white", font=("Helvetica", 10))
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Debes seleccionar un rango de fechas."))

    root.mainloop()

    if start_date is None or end_date is None:
        raise ValueError("No s'ha seleccionat cap rang de dates")
    
    return start_date, end_date

def show_date_range(selected_date, to_date):
    # Aquí deberías poner la lógica para seleccionar las fechas
    get_date_range(selected_date, to_date)
    return selected_date, to_date

def exit_program(root):
    if messagebox.askokcancel("Sortir", "Està segur que voleu sortir?"):
        root.destroy()
        sys.exit()
        
def get_date(val):
    """Muestra un calendario para que el usuario seleccione una fecha."""
    root = Tk()
    root.geometry("300x350")
    if val == 0:
        root.title("Seleccionar data d'inici")
    else:
        root.title("Seleccionar data final")

    
    def get_selected_date():
        """Obtiene la fecha seleccionada por el usuario y cierra la ventana."""
        nonlocal selected_date
        selected_date = cal.selection_get()
        root.destroy()

    def cancel():
        """Cierra la ventana sin seleccionar una fecha."""
        nonlocal selected_date
        selected_date = None
        root.destroy()
        exit()

     #Poner un titulo en la ventana
    if val == 0:
        title_label = Label(root, text="Seleccionar data d'inici", font=("Helvetica", 16), fg="black", bg="white")
        title_label.pack(pady=10)
    else:
        title_label = Label(root, text="Seleccionar data final", font=("Helvetica", 16), fg="black", bg="white")
        title_label.pack(pady=10)

    selected_date = None
    

    cal = Calendar(root, selectmode="day", year=now.year , month=now.month, day=now.day,maxdate=now)
    cal.pack(pady=20)


    btn_ok = Button(root, text="OK", command=get_selected_date)
    btn_ok.pack(side="left", pady=10, padx=10)

    btn_cancel = Button(root, text="CANCEL·LAR", command=cancel)
    btn_cancel.pack(side="right", pady=10, padx=10)

    root.protocol("WM_DELETE_WINDOW", lambda: messagebox.showerror("Error", "Heu de seleccionar una data"))

    root.mainloop()

    if selected_date is None:
        raise ValueError("No s'ha seleccionat cap data")
    return selected_date


def main():
    global df
    global generar_desplegaments_PRO
    print("###############################################")
    print("##                                           ##")
    print("##  Programa de generació d'informes de      ##")
    print("##  gestió de desplegaments d'aplicacions    ##")
    print("##                                           ##")
    print("###############################################")
    print()
    print("Sisplau, seleccioni les dates del rang de temps que vol consultar:")

    # Obtener la fecha seleccionada por el usuario y calcular la fecha inicial restando 4 días
    selected_date = get_date(0)
    print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")

    to_date = get_date(1)
    print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")

    #Espera 1 segon per a que es mostri la data final
    time.sleep(1)

    print("Opcions:")
    print("1. Revisar les dates")
    print("2. Generar informe setmanal")
    print("3. Generar informe de desplegaments de PRO")
    print("4. Sortir")
    print()

    # creamos un bucle para que el usuario pueda elegir una opcion hasta que la opcion sea correcta
    while True:
        try:
            # pedimos al usuario que introduzca una opcion
            opcion = int(input("Introdueix una opció: "))
            # si la opcion es correcta salimos del bucle
            if opcion == 1:
                selected_date = get_date(0)
                print(f"Data d'inici seleccionada: -------------------> {selected_date.strftime('%d/%m/%Y')}")

                to_date = get_date(1)
                print(f"Data final seleccionada: -------------------> {to_date.strftime('%d/%m/%Y')}")
                time.sleep(1)
                continue
            elif opcion == 2:
                break
            elif opcion == 3:
                generar_desplegaments_PRO = True
                break
            elif opcion == 4:
                exit()
            else:
                print("Opció incorrecta")
                continue
        except ValueError:
            print("Opció incorrecta")
            continue

    from_date_week = selected_date - datetime.timedelta(days=7)
    
    # Conectar con Outlook y obtener la carpeta de la bandeja de entrada
    print("Connectant al vostre compte d'Outlook...")
    outlook = connect_to_outlook()
    inbox = get_inbox_folder(outlook, "gestioversions@bcn.cat","Bandeja de entrada")
   

    # Obtener los mensajes de la bandeja de entrada en el rango de fechas especificado
    print(f"Obtenint els missatges de {from_date_week.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    print()
    print("Si us plau espera un moment...")
    print()
    messages = get_inbox_messages(inbox, from_date_week, to_date)

    
    # Extraer información relevante de los correos electrónicos y clasificarlos en el dataframe
    print(f"Realitzar l'extracció dels mails de {selected_date.strftime('%d/%m/%Y')} fins {to_date.strftime('%d/%m/%Y')}")
    emails, datas = extract_emails(messages, selected_date, to_date)

    df,df_despliegues = classify_message_and_inter_into_dataframe(emails, df)

   
    print("Extracció realitzada correctament, se ha creat al dataset !!")

    #Si el dataframe esta vacio, mostrar un mensaje de error y salir del programa
    if df.empty:
        messagebox.showerror("Error", "No s'han trobat correus electrònics al rang de dates especificat")
        sys.exit()
    
    # Agregar la columna 'datas' en el orden de la lista, pero solo a aquellas filas del dataframe que no tengan fecha
    df.loc[df['Data'] == '', 'Data'] = datas

   
    # Pasar el dataframe a un archivo de Excel y guardar en disco
    if generar_desplegaments_PRO:
        pass_df_despliegues_to_excel(df_despliegues, selected_date, to_date)
    else:
        pass_df_to_excel(df, selected_date, to_date)

    print("Fitxer Excel generat correctament !!")
    print()
    print("###############################################")
    print("##  fi del programa                          ##")
    print("###############################################")
    

if __name__ == "__main__":
    try:
        main()
        input("Premeu una tecla per a continuar...")
    except Exception as e:
        #mostrar la linea de error
        messagebox.showerror("Error", f"Ha ocurrido un error: {e}")
        sys.exit()
   

    
 
        